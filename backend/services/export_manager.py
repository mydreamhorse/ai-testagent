"""
导出管理器
负责将报告导出为不同格式（PDF、Excel、HTML等）
"""

import os
import io
import json
import uuid
from typing import Dict, List, Optional, Any, Union, BinaryIO
from datetime import datetime
from enum import Enum
from dataclasses import dataclass
from pathlib import Path
import asyncio
import queue
import threading
import logging

# PDF generation
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError):
    WEASYPRINT_AVAILABLE = False

# Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from sqlalchemy.orm import Session
from backend.models import Report, ReportTemplate
from backend.database import get_db

logger = logging.getLogger(__name__)


class ExportFormat(Enum):
    """导出格式枚举"""
    PDF = "pdf"
    EXCEL = "excel"
    HTML = "html"
    JSON = "json"
    CSV = "csv"


class ExportStatus(Enum):
    """导出状态枚举"""
    PENDING = "pending"
    IN_PROGRESS = "in_progress"
    COMPLETED = "completed"
    FAILED = "failed"


@dataclass
class ExportTask:
    """导出任务"""
    id: str
    report_id: int
    format: ExportFormat
    options: Dict[str, Any]
    status: ExportStatus
    created_at: datetime
    started_at: Optional[datetime] = None
    completed_at: Optional[datetime] = None
    file_path: Optional[str] = None
    error_message: Optional[str] = None
    progress: float = 0.0


@dataclass
class ExportResult:
    """导出结果"""
    success: bool
    file_path: Optional[str] = None
    file_size: Optional[int] = None
    error_message: Optional[str] = None
    metadata: Optional[Dict[str, Any]] = None


class PDFExporter:
    """PDF导出器"""
    
    def __init__(self):
        if not WEASYPRINT_AVAILABLE:
            raise ImportError("WeasyPrint is required for PDF export. Install with: pip install weasyprint")
    
    def export(self, content: str, output_path: str, options: Dict[str, Any] = None) -> ExportResult:
        """导出PDF"""
        try:
            options = options or {}
            
            # 准备HTML内容
            html_content = self._prepare_html_content(content, options)
            
            # 准备CSS样式
            css_content = self._prepare_css_styles(options)
            
            # 生成PDF
            html_doc = HTML(string=html_content)
            css_doc = CSS(string=css_content) if css_content else None
            
            if css_doc:
                html_doc.write_pdf(output_path, stylesheets=[css_doc])
            else:
                html_doc.write_pdf(output_path)
            
            # 获取文件大小
            file_size = os.path.getsize(output_path)
            
            return ExportResult(
                success=True,
                file_path=output_path,
                file_size=file_size,
                metadata={
                    'format': 'pdf',
                    'pages': self._count_pdf_pages(output_path),
                    'generated_at': datetime.utcnow().isoformat()
                }
            )
            
        except Exception as e:
            logger.error(f"PDF export failed: {str(e)}")
            return ExportResult(
                success=False,
                error_message=f"PDF导出失败: {str(e)}"
            )
    
    def _prepare_html_content(self, content: str, options: Dict[str, Any]) -> str:
        """准备HTML内容"""
        # 如果内容已经是完整的HTML，直接返回
        if content.strip().startswith('<!DOCTYPE') or content.strip().startswith('<html'):
            return content
        
        # 否则包装成完整的HTML文档
        title = options.get('title', '报告')
        
        html_template = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
</head>
<body>
    <div class="report-content">
        {content}
    </div>
</body>
</html>
        """.strip()
        
        return html_template
    
    def _prepare_css_styles(self, options: Dict[str, Any]) -> str:
        """准备CSS样式"""
        default_css = """
        @page {
            size: A4;
            margin: 2cm;
        }
        
        body {
            font-family: 'SimSun', serif;
            font-size: 12pt;
            line-height: 1.6;
            color: #333;
        }
        
        .report-content {
            max-width: 100%;
        }
        
        h1, h2, h3, h4, h5, h6 {
            color: #2c3e50;
            margin-top: 1.5em;
            margin-bottom: 0.5em;
        }
        
        h1 { font-size: 18pt; }
        h2 { font-size: 16pt; }
        h3 { font-size: 14pt; }
        
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 1em 0;
        }
        
        th, td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }
        
        th {
            background-color: #f5f5f5;
            font-weight: bold;
        }
        
        .badge {
            display: inline-block;
            padding: 0.25em 0.4em;
            font-size: 0.875em;
            font-weight: 700;
            line-height: 1;
            text-align: center;
            white-space: nowrap;
            vertical-align: baseline;
            border-radius: 0.25rem;
        }
        
        .badge-success { background-color: #28a745; color: white; }
        .badge-danger { background-color: #dc3545; color: white; }
        .badge-warning { background-color: #ffc107; color: black; }
        .badge-secondary { background-color: #6c757d; color: white; }
        .badge-primary { background-color: #007bff; color: white; }
        
        .page-break {
            page-break-before: always;
        }
        """
        
        # 合并自定义CSS
        custom_css = options.get('css', '')
        if custom_css:
            default_css += f"\n\n/* Custom CSS */\n{custom_css}"
        
        return default_css
    
    def _count_pdf_pages(self, pdf_path: str) -> int:
        """计算PDF页数（简化实现）"""
        try:
            # 这里可以使用PyPDF2或其他库来准确计算页数
            # 目前返回估算值
            file_size = os.path.getsize(pdf_path)
            # 粗略估算：每页约50KB
            estimated_pages = max(1, file_size // 50000)
            return estimated_pages
        except:
            return 1


class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    def export(self, data: Dict[str, Any], output_path: str, options: Dict[str, Any] = None) -> ExportResult:
        """导出Excel"""
        try:
            options = options or {}
            
            # 创建工作簿
            workbook = openpyxl.Workbook()
            
            # 删除默认工作表
            workbook.remove(workbook.active)
            
            # 根据数据类型创建不同的工作表
            self._create_summary_sheet(workbook, data, options)
            
            if 'test_cases' in data:
                self._create_test_cases_sheet(workbook, data['test_cases'], options)
            
            if 'defects' in data:
                self._create_defects_sheet(workbook, data['defects'], options)
            
            if 'coverage_data' in data:
                self._create_coverage_sheet(workbook, data['coverage_data'], options)
            
            # 保存文件
            workbook.save(output_path)
            
            # 获取文件大小
            file_size = os.path.getsize(output_path)
            
            return ExportResult(
                success=True,
                file_path=output_path,
                file_size=file_size,
                metadata={
                    'format': 'excel',
                    'sheets': len(workbook.worksheets),
                    'generated_at': datetime.utcnow().isoformat()
                }
            )
            
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}")
            return ExportResult(
                success=False,
                error_message=f"Excel导出失败: {str(e)}"
            )
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook, data: Dict[str, Any], options: Dict[str, Any]):
        """创建摘要工作表"""
        sheet = workbook.create_sheet("摘要", 0)
        
        # 设置标题
        title = options.get('title', '测试报告摘要')
        sheet['A1'] = title
        sheet['A1'].font = Font(size=16, bold=True)
        sheet.merge_cells('A1:D1')
        
        # 基本信息
        row = 3
        basic_info = [
            ('报告生成时间', data.get('report_generated_at', datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'))),
            ('测试用例总数', data.get('total_test_cases', 0)),
            ('通过数量', data.get('passed_count', 0)),
            ('失败数量', data.get('failed_count', 0)),
            ('通过率', f"{data.get('pass_rate', 0) * 100:.1f}%" if data.get('pass_rate') else 'N/A')
        ]
        
        for label, value in basic_info:
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # 设置列宽
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 30
    
    def _create_test_cases_sheet(self, workbook: openpyxl.Workbook, test_cases: List[Dict], options: Dict[str, Any]):
        """创建测试用例工作表"""
        sheet = workbook.create_sheet("测试用例")
        
        # 表头
        headers = ['ID', '标题', '类型', '状态', '优先级', '创建时间']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 数据行
        for row, test_case in enumerate(test_cases, 2):
            sheet.cell(row=row, column=1, value=test_case.get('id', ''))
            sheet.cell(row=row, column=2, value=test_case.get('title', ''))
            sheet.cell(row=row, column=3, value=test_case.get('test_type', ''))
            sheet.cell(row=row, column=4, value=test_case.get('status', ''))
            sheet.cell(row=row, column=5, value=test_case.get('priority', ''))
            sheet.cell(row=row, column=6, value=test_case.get('created_at', ''))
        
        # 自动调整列宽
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_defects_sheet(self, workbook: openpyxl.Workbook, defects: List[Dict], options: Dict[str, Any]):
        """创建缺陷工作表"""
        sheet = workbook.create_sheet("缺陷")
        
        # 表头
        headers = ['ID', '类型', '严重程度', '描述', '状态', '检测时间']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 数据行
        for row, defect in enumerate(defects, 2):
            sheet.cell(row=row, column=1, value=defect.get('id', ''))
            sheet.cell(row=row, column=2, value=defect.get('defect_type', ''))
            sheet.cell(row=row, column=3, value=defect.get('severity', ''))
            sheet.cell(row=row, column=4, value=defect.get('description', ''))
            sheet.cell(row=row, column=5, value=defect.get('status', ''))
            sheet.cell(row=row, column=6, value=defect.get('detected_at', ''))
        
        # 自动调整列宽
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_coverage_sheet(self, workbook: openpyxl.Workbook, coverage_data: List[Dict], options: Dict[str, Any]):
        """创建覆盖率工作表"""
        sheet = workbook.create_sheet("覆盖率")
        
        # 表头
        headers = ['功能模块', '覆盖率', '已覆盖测试用例', '总测试用例', '分析日期']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 数据行
        for row, coverage in enumerate(coverage_data, 2):
            sheet.cell(row=row, column=1, value=coverage.get('function_module', ''))
            sheet.cell(row=row, column=2, value=f"{coverage.get('coverage_percentage', 0):.1f}%")
            sheet.cell(row=row, column=3, value=coverage.get('covered_test_cases', 0))
            sheet.cell(row=row, column=4, value=coverage.get('total_test_cases', 0))
            sheet.cell(row=row, column=5, value=coverage.get('analysis_date', ''))
        
        # 自动调整列宽
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[column_letter].width = adjusted_width


class HTMLExporter:
    """HTML导出器"""
    
    def export(self, content: str, output_path: str, options: Dict[str, Any] = None) -> ExportResult:
        """导出HTML"""
        try:
            options = options or {}
            
            # 准备完整的HTML内容
            html_content = self._prepare_html_content(content, options)
            
            # 写入文件
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            # 获取文件大小
            file_size = os.path.getsize(output_path)
            
            return ExportResult(
                success=True,
                file_path=output_path,
                file_size=file_size,
                metadata={
                    'format': 'html',
                    'encoding': 'utf-8',
                    'generated_at': datetime.utcnow().isoformat()
                }
            )
            
        except Exception as e:
            logger.error(f"HTML export failed: {str(e)}")
            return ExportResult(
                success=False,
                error_message=f"HTML导出失败: {str(e)}"
            )
    
    def _prepare_html_content(self, content: str, options: Dict[str, Any]) -> str:
        """准备HTML内容"""
        # 如果内容已经是完整的HTML，直接返回
        if content.strip().startswith('<!DOCTYPE') or content.strip().startswith('<html'):
            return content
        
        title = options.get('title', '报告')
        css_styles = self._get_css_styles(options)
        
        html_template = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
{css_styles}
    </style>
</head>
<body>
    <div class="container">
        <div class="report-content">
{content}
        </div>
        <div class="footer">
            <p>报告生成时间: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>"""
        
        return html_template
    
    def _get_css_styles(self, options: Dict[str, Any]) -> str:
        """获取CSS样式"""
        default_css = """
        body {
            font-family: 'Microsoft YaHei', 'SimSun', sans-serif;
            line-height: 1.6;
            color: #333;
            margin: 0;
            padding: 0;
            background-color: #f5f5f5;
        }
        
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            box-shadow: 0 0 10px rgba(0,0,0,0.1);
            min-height: 100vh;
        }
        
        .report-content {
            padding: 40px;
        }
        
        h1, h2, h3, h4, h5, h6 {
            color: #2c3e50;
            margin-top: 2em;
            margin-bottom: 1em;
        }
        
        h1 { font-size: 2.5em; border-bottom: 3px solid #3498db; padding-bottom: 10px; }
        h2 { font-size: 2em; border-bottom: 2px solid #e74c3c; padding-bottom: 8px; }
        h3 { font-size: 1.5em; border-bottom: 1px solid #95a5a6; padding-bottom: 5px; }
        
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        }
        
        th, td {
            border: 1px solid #ddd;
            padding: 12px;
            text-align: left;
        }
        
        th {
            background-color: #3498db;
            color: white;
            font-weight: bold;
        }
        
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        
        tr:hover {
            background-color: #f5f5f5;
        }
        
        .badge {
            display: inline-block;
            padding: 0.25em 0.6em;
            font-size: 0.875em;
            font-weight: 700;
            line-height: 1;
            text-align: center;
            white-space: nowrap;
            vertical-align: baseline;
            border-radius: 0.375rem;
        }
        
        .badge-success { background-color: #28a745; color: white; }
        .badge-danger { background-color: #dc3545; color: white; }
        .badge-warning { background-color: #ffc107; color: #212529; }
        .badge-secondary { background-color: #6c757d; color: white; }
        .badge-primary { background-color: #007bff; color: white; }
        
        .footer {
            background-color: #f8f9fa;
            padding: 20px 40px;
            border-top: 1px solid #dee2e6;
            text-align: center;
            color: #6c757d;
            font-size: 0.9em;
        }
        
        .chart-container {
            margin: 20px 0;
            text-align: center;
        }
        
        .metric-card {
            display: inline-block;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            margin: 10px;
            border-radius: 10px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.2);
            min-width: 150px;
        }
        
        .metric-value {
            font-size: 2em;
            font-weight: bold;
            margin-bottom: 5px;
        }
        
        .metric-label {
            font-size: 0.9em;
            opacity: 0.9;
        }
        """
        
        # 合并自定义CSS
        custom_css = options.get('css', '')
        if custom_css:
            default_css += f"\n\n/* Custom CSS */\n{custom_css}"
        
        return default_css


class ExportTaskQueue:
    """导出任务队列"""
    
    def __init__(self, max_workers: int = 3):
        self.task_queue = queue.Queue()
        self.tasks = {}  # task_id -> ExportTask
        self.max_workers = max_workers
        self.workers = []
        self.running = False
        
        # 创建导出器实例
        self.exporters = {
            ExportFormat.PDF: PDFExporter() if WEASYPRINT_AVAILABLE else None,
            ExportFormat.EXCEL: ExcelExporter() if OPENPYXL_AVAILABLE else None,
            ExportFormat.HTML: HTMLExporter(),
        }
    
    def start(self):
        """启动任务队列"""
        if self.running:
            return
        
        self.running = True
        for i in range(self.max_workers):
            worker = threading.Thread(target=self._worker, name=f"ExportWorker-{i}")
            worker.daemon = True
            worker.start()
            self.workers.append(worker)
        
        logger.info(f"Export task queue started with {self.max_workers} workers")
    
    def stop(self):
        """停止任务队列"""
        self.running = False
        
        # 向队列中添加停止信号
        for _ in range(self.max_workers):
            self.task_queue.put(None)
        
        # 等待所有工作线程结束
        for worker in self.workers:
            worker.join(timeout=5)
        
        self.workers.clear()
        logger.info("Export task queue stopped")
    
    def add_task(self, task: ExportTask) -> str:
        """添加导出任务"""
        self.tasks[task.id] = task
        self.task_queue.put(task.id)
        logger.info(f"Added export task {task.id} to queue")
        return task.id
    
    def get_task_status(self, task_id: str) -> Optional[ExportTask]:
        """获取任务状态"""
        return self.tasks.get(task_id)
    
    def _worker(self):
        """工作线程"""
        while self.running:
            try:
                task_id = self.task_queue.get(timeout=1)
                if task_id is None:  # 停止信号
                    break
                
                task = self.tasks.get(task_id)
                if not task:
                    continue
                
                self._process_task(task)
                
            except queue.Empty:
                continue
            except Exception as e:
                logger.error(f"Export worker error: {str(e)}")
    
    def _process_task(self, task: ExportTask):
        """处理导出任务"""
        try:
            task.status = ExportStatus.IN_PROGRESS
            task.started_at = datetime.utcnow()
            task.progress = 0.1
            
            logger.info(f"Processing export task {task.id}")
            
            # 获取对应的导出器
            exporter = self.exporters.get(task.format)
            if not exporter:
                raise ValueError(f"Unsupported export format: {task.format}")
            
            task.progress = 0.3
            
            # 执行导出
            result = self._execute_export(task, exporter)
            
            task.progress = 0.9
            
            if result.success:
                task.status = ExportStatus.COMPLETED
                task.file_path = result.file_path
                task.progress = 1.0
                logger.info(f"Export task {task.id} completed successfully")
            else:
                task.status = ExportStatus.FAILED
                task.error_message = result.error_message
                logger.error(f"Export task {task.id} failed: {result.error_message}")
            
            task.completed_at = datetime.utcnow()
            
        except Exception as e:
            task.status = ExportStatus.FAILED
            task.error_message = str(e)
            task.completed_at = datetime.utcnow()
            logger.error(f"Export task {task.id} failed with exception: {str(e)}")
    
    def _execute_export(self, task: ExportTask, exporter) -> ExportResult:
        """执行具体的导出操作"""
        # 这里需要根据task.report_id获取报告数据
        # 简化实现，实际应该从数据库获取
        
        if task.format == ExportFormat.HTML:
            # HTML导出
            content = task.options.get('content', '<h1>测试报告</h1><p>报告内容</p>')
            return exporter.export(content, task.options.get('output_path'), task.options)
        
        elif task.format == ExportFormat.PDF:
            # PDF导出
            content = task.options.get('content', '<h1>测试报告</h1><p>报告内容</p>')
            return exporter.export(content, task.options.get('output_path'), task.options)
        
        elif task.format == ExportFormat.EXCEL:
            # Excel导出
            data = task.options.get('data', {})
            return exporter.export(data, task.options.get('output_path'), task.options)
        
        else:
            return ExportResult(success=False, error_message=f"Unsupported format: {task.format}")


class ExportManager:
    """导出管理器"""
    
    def __init__(self, db: Session = None, upload_dir: str = "uploads/exports"):
        self.db = db or next(get_db())
        self.upload_dir = Path(upload_dir)
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        
        # 初始化任务队列
        self.task_queue = ExportTaskQueue()
        self.task_queue.start()
    
    def __del__(self):
        """析构函数"""
        if hasattr(self, 'task_queue'):
            self.task_queue.stop()
    
    def export_report(self, report_id: int, format: ExportFormat, options: Dict[str, Any] = None) -> str:
        """导出报告"""
        options = options or {}
        
        # 生成任务ID和文件路径
        task_id = str(uuid.uuid4())
        filename = self._generate_filename(report_id, format, options)
        output_path = str(self.upload_dir / filename)
        
        # 准备任务选项
        task_options = options.copy()
        task_options['output_path'] = output_path
        
        # 获取报告数据
        report_data = self._get_report_data(report_id)
        if format == ExportFormat.EXCEL:
            task_options['data'] = report_data
        else:
            task_options['content'] = self._render_report_content(report_data, options)
        
        # 创建导出任务
        task = ExportTask(
            id=task_id,
            report_id=report_id,
            format=format,
            options=task_options,
            status=ExportStatus.PENDING,
            created_at=datetime.utcnow()
        )
        
        # 添加到队列
        self.task_queue.add_task(task)
        
        return task_id
    
    def get_export_status(self, task_id: str) -> Optional[ExportTask]:
        """获取导出状态"""
        return self.task_queue.get_task_status(task_id)
    
    def get_supported_formats(self) -> List[str]:
        """获取支持的导出格式"""
        formats = [ExportFormat.HTML.value]
        
        if WEASYPRINT_AVAILABLE:
            formats.append(ExportFormat.PDF.value)
        
        if OPENPYXL_AVAILABLE:
            formats.append(ExportFormat.EXCEL.value)
        
        return formats
    
    def _generate_filename(self, report_id: int, format: ExportFormat, options: Dict[str, Any]) -> str:
        """生成文件名"""
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        prefix = options.get('filename_prefix', f'report_{report_id}')
        extension = format.value
        
        if format == ExportFormat.EXCEL:
            extension = 'xlsx'
        
        return f"{prefix}_{timestamp}.{extension}"
    
    def _get_report_data(self, report_id: int) -> Dict[str, Any]:
        """获取报告数据"""
        # 从数据库获取报告数据
        report = self.db.query(Report).filter(Report.id == report_id).first()
        if not report:
            return {}
        
        # 返回报告数据
        return {
            'id': report.id,
            'title': report.title,
            'report_type': report.report_type,
            'generation_time': report.generation_time,
            'report_data': report.report_data or {},
            'report_generated_at': report.generation_time.strftime('%Y-%m-%d %H:%M:%S') if report.generation_time else None
        }
    
    def _render_report_content(self, report_data: Dict[str, Any], options: Dict[str, Any]) -> str:
        """渲染报告内容"""
        # 简化的内容渲染
        title = report_data.get('title', '测试报告')
        report_type = report_data.get('report_type', 'unknown')
        generation_time = report_data.get('report_generated_at', 'N/A')
        
        content = f"""
<h1>{title}</h1>

<h2>基本信息</h2>
<ul>
    <li><strong>报告类型:</strong> {report_type}</li>
    <li><strong>生成时间:</strong> {generation_time}</li>
</ul>

<h2>报告内容</h2>
<p>这是一个示例报告内容。在实际应用中，这里应该包含详细的测试数据和分析结果。</p>
        """.strip()
        
        return content
    
    def cleanup_old_files(self, days: int = 7):
        """清理旧的导出文件"""
        try:
            cutoff_time = datetime.utcnow() - timedelta(days=days)
            
            for file_path in self.upload_dir.glob("*"):
                if file_path.is_file():
                    file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                    if file_time < cutoff_time:
                        file_path.unlink()
                        logger.info(f"Cleaned up old export file: {file_path}")
        
        except Exception as e:
            logger.error(f"Error cleaning up old files: {str(e)}")


# 全局导出管理器实例
_export_manager = None


def get_export_manager() -> ExportManager:
    """获取导出管理器实例"""
    global _export_manager
    if _export_manager is None:
        _export_manager = ExportManager()
    return _export_manager